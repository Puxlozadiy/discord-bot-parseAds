# coding=utf-8
from discord import client
from discord.flags import Intents
from discord.message import MessageReference
from openpyxl import load_workbook, Workbook
wb = load_workbook('Analusis.xlsx')
ws = wb.active
import discord
from discord.ext import commands, tasks
 
bot = commands.Bot(command_prefix='!', self_bot=True)
target_channel_id = 886915166120345670
target_channel_id1 = 652634813576118335


@bot.event
async def on_message(message):
    if message.channel.id == target_channel_id1:
        async for message in message.channel.history(limit=1):
            ifContainsAndStore(message)

def ifContainsAndStore(string):
    if string.content.find("родам") >= 0:
        if string.content.find("evo 6") >= 0 or string.content.find("Evo 6") >= 0 or string.content.find("евик 6") >= 0 or string.content.find("Evo6") >= 0 or string.content.find("EVO 6") >= 0 or string.content.find("Эвик 6") >= 0:
            if string.content.find("evo 6") >= 0:
                storeData(string.content, 3, "Продаётся evo 6", "B", string.content.find("evo 6"))
            elif string.content.find("evo 6") >= 0:
                storeData(string.content, 3, "Продаётся evo 6", "B", string.content.find("Evo 6"))
            elif string.content.find("evo 6") >= 0:
                storeData(string.content, 3, "Продаётся evo 6", "B", string.content.find("евик 6"))
            elif string.content.find("evo 6") >= 0:
                storeData(string.content, 3, "Продаётся evo 6", "B", string.content.find("Evo6"))
            elif string.content.find("evo 6") >= 0:
                storeData(string.content, 3, "Продаётся evo 6", "B", string.content.find("EVO 6"))
            elif string.content.find("evo 6") >= 0:
                storeData(string.content, 3, "Продаётся evo 6", "B", string.content.find("Эвик 6"))
        elif string.content.find("350z") >= 0:
            if string.content.find("350z") >= 0:
                storeData(string.content, 4, "Продаётся 350z", "C", string.content.find("350z"))
        elif string.content.find("s15") >= 0:
            if string.content.find("s15") >= 0:
                storeData(string.content, 5, "Продаётся silvia s15", "D", string.content.find("s15"))
        elif string.content.find("FX50") >= 0 or string.content.find("финик") >= 0:
            if string.content.find("FX50") >= 0:
                storeData(string.content, 6, "Продаётся Infinity", "F", string.content.find("FX50"))
            elif string.content.find("FX50") >= 0:
                storeData(string.content, 6, "Продаётся Infinity", "F", string.content.find("финик"))
        elif string.content.find("Supra") >= 0 or string.content.find("супру") >= 0:
            if string.content.find("Supra") >= 0:
                storeData(string.content, 7, "Продаётся Toyota Supra", "F", string.content.find("Supra"))
            elif string.content.find("супру") >= 0:
                storeData(string.content, 7, "Продаётся Toyota Supra", "F", string.content.find("супру"))
        elif string.content.find("WRX STI") >= 0 :
            if string.content.find("WRX STI") >= 0:
                storeData(string.content, 8, "Продаётся Subaru WRX STI", "G", string.content.find("WRX STI"))
        elif string.content.find("Stinger") >= 0 or string.content.find("стингер") >= 0 or string.content.find("Киа") >= 0:
            if string.content.find("Stinger") >= 0:
                storeData(string.content, 9, "Продаётся KIA Stinger", "H", string.content.find("Stinger"))
            elif string.content.find("стингер") >= 0:
                storeData(string.content, 9, "Продаётся KIA Stinger", "H", string.content.find("стингер"))
            elif string.content.find("Киа") >= 0:
                storeData(string.content, 9, "Продаётся KIA Stinger", "H", string.content.find("Киа"))
        elif string.content.find("R34") >= 0 or string.content.find("скайлик") >= 0 or string.content.find("34") >= 0:
            if string.content.find("R34") >= 0:
                storeData(string.content, 10, "Продаётся Skyline R34", "I", string.content.find("R34"))
            elif string.content.find("скайлик 34") >= 0:
                storeData(string.content, 10, "Продаётся Skyline R34", "I", string.content.find("скайлик 34"))
            elif string.content.find("34") >= 0:
                storeData(string.content, 10, "Продаётся Skyline R34", "I", string.content.find("34"))
        elif string.content.find("E70") >= 0 and string.content.find("X5") >= 0 or string.content.find("e70") >= 0 and string.content.find("x5") >= 0 or string.content.find("E70 X5M") >= 0 or string.content.find("X5 e70") >= 0:
            if string.content.find("E70") >= 0:
                storeData(string.content, 11, "Продаётся BMW x5 e70", "J", string.content.find("E70"))
            elif string.content.find("e70") >= 0:
                storeData(string.content, 11, "Продаётся BMW x5 e70", "J", string.content.find("e70"))
            elif string.content.find("x5") >= 0:
                storeData(string.content, 11, "Продаётся BMW x5 e70", "J", string.content.find("x5"))
            elif string.content.find("E70 X5M") >= 0:
                storeData(string.content, 11, "Продаётся BMW x5 e70", "J", string.content.find("E70 X5M"))
            elif string.content.find("X5 e70") >= 0:
                storeData(string.content, 11, "Продаётся BMW x5 e70", "J", string.content.find("X5 e70"))
        elif string.content.find("Alfa") >= 0 or string.content.find("альфу") >= 0:
            if string.content.find("Alfa") >= 0:
                storeData(string.content, 12, "Продаётся Alfa Romeo", "K", string.content.find("Alfa"))
            elif string.content.find("альфу") >= 0:
                storeData(string.content, 12, "Продаётся Alfa Romeo", "K", string.content.find("альфу"))
        elif string.content.find("Raptor") >= 0 or string.content.find("раптор") >= 0 or string.content.find("Раптор") >= 0 or string.content.find("raptor") >= 0:
            if string.content.find("Raptor") >= 0:
                storeData(string.content, 13, "Продаётся Ford Raptor", "L", string.content.find("Raptor"))
            elif string.content.find("раптор") >= 0:
                storeData(string.content, 13, "Продаётся Ford Raptor", "L", string.content.find("раптор"))
            elif string.content.find("Раптор") >= 0:
                storeData(string.content, 13, "Продаётся Ford Raptor", "L", string.content.find("Раптор"))
            elif string.content.find("raptor") >= 0:
                storeData(string.content, 13, "Продаётся Ford Raptor", "L", string.content.find("raptor"))
        elif string.content.find("GLS63") >= 0 or string.content.find("GLS") >= 0 or string.content.find("глс") >= 0 or string.content.find("ГЛС") >= 0:
            if string.content.find("GLS63") >= 0:
                storeData(string.content, 14, "Продаётся Mercedes GLS63", "M", string.content.find("GLS63"))
            elif string.content.find("GLS") >= 0:
                storeData(string.content, 14, "Продаётся Mercedes GLS63", "M", string.content.find("GLS"))
            elif string.content.find("глс") >= 0:
                storeData(string.content, 14, "Продаётся Mercedes GLS63", "M", string.content.find("глс"))
            elif string.content.find("ГЛС") >= 0:
                storeData(string.content, 14, "Продаётся Mercedes GLS63", "M", string.content.find("ГЛС"))
        elif string.content.find("Mustang 2019") >= 0 :
            if string.content.find("Mustang 2019") >= 0:
                storeData(string.content, 15, "Продаётся Ford Mustang 2019", "N", string.content.find("Mustang 2019"))
        elif string.content.find("M5F90") >= 0:
            if string.content.find("M5F90") >= 0:
                storeData(string.content, 16, "Продаётся BMW M5F90", "O", string.content.find("M5F90"))
        elif string.content.find("LX570") >= 0 or string.content.find("лексус") >= 0 or string.content.find("Лексус") >= 0:
            if string.content.find("LX570") >= 0:
                storeData(string.content, 17, "Продаётся Lexus LX570", "P", string.content.find("LX570"))
            elif string.content.find("лексус") >= 0:
                storeData(string.content, 17, "Продаётся Lexus LX570", "P", string.content.find("лексус"))
            elif string.content.find("Лексус") >= 0:
                storeData(string.content, 17, "Продаётся Lexus LX570", "P", string.content.find("Лексус"))
        elif string.content.find("Model S") >= 0 :
            if string.content.find("Model S") >= 0:
                storeData(string.content, 18, "Продаётся Tesla Model S", "Q", string.content.find("Model S"))
        elif string.content.find("G63") >= 0 or string.content.find("гелик") >= 0 or string.content.find("Гелик") >= 0:
            if string.content.find("G63") >= 0:
                storeData(string.content, 19, "Продаётся Mercedes G63", "R", string.content.find("G63"))
            elif string.content.find("гелик") >= 0:
                storeData(string.content, 19, "Продаётся Mercedes G63", "R", string.content.find("гелик"))
            elif string.content.find("Гелик") >= 0:
                storeData(string.content, 19, "Продаётся Mercedes G63", "R", string.content.find("Гелик"))
        elif string.content.find("RS7") >= 0 :
            if string.content.find("RS7") >= 0:
                storeData(string.content, 20, "Продаётся Audi RS7", "S", string.content.find("RS7"))
        elif string.content.find("X6 M") >= 0 or string.content.find("X6M") >= 0 or string.content.find("x6m") >= 0 or string.content.find("X6 m") >= 0 or string.content.find("х6м") >= 0:
            if string.content.find("X6 M") >= 0:
                storeData(string.content, 21, "Продаётся BMW X6 M", "T", string.content.find("X6 M"))
            elif string.content.find("X6 M") >= 0:
                storeData(string.content, 21, "Продаётся BMW X6 M", "T", string.content.find("X6M"))
            elif string.content.find("X6 M") >= 0:
                storeData(string.content, 21, "Продаётся BMW X6 M", "T", string.content.find("x6m"))
            elif string.content.find("X6 M") >= 0:
                storeData(string.content, 21, "Продаётся BMW X6 M", "T", string.content.find("X6 m"))
            elif string.content.find("X6 M") >= 0:
                storeData(string.content, 21, "Продаётся BMW X6 M", "T", string.content.find("х6м"))

def storeData(string, int, comment, letter, position):
    converted = convertToMoney(string, position)
    if converted != "Цена договорная!" and converted != "Скорее всего это обмен!" and converted != None:
        if converted != 0:
            x = ws[f'A{int}'].value
            ws[f'{letter}{x}'].value = converted
            ws[f'A{int}'].value = (x + 1)
            wb.save('Analusis.xlsx')
            print(comment, "за", converted)
        else:
            print("Ошибка в получении цены или цена договорная!")
    else:
        print("Цена договорная и она не записана в таблицу!")  

def convertToMoney(string, position):
    bool = 0
    millions = 0
    thousands = 0
    hundreds = 0
    block = 0
    count = 0
    count1 = 0
    for lines in string.splitlines():
        if lines.find('догов') >= 0 or string.find('Догов') >= 0 or string.find('в лс') >= 0:
            return "Цена договорная!"
        if lines.find('бмен') == -1 or string.find('Обмен') == -1 or string.find('обмен') == -1:
            if lines.find('гос') == -1 and lines.find('Гос') == -1 and lines.find('Гос. цена') == -1:
                for str in lines.split(): # !!!!ИДЕЯ: можно не пускать цикл, пока нужное название машины не будет найдено
                    if string.find(str) > position:
                        if str.find('моей') >= 0 or str.find('дп') >= 0 or str.find('Дп') >= 0 or str.find('доплата') >= 0 or str.find('Доплата') >= 0:
                            if count1 == 0:
                                    count1 = 1
                                    continue
                        elif count1 == 1:
                            count1 = 0
                            continue
                        if count1 == 0:
                            if str.find("кк") >= 0 or str.find(",") >= 0 or str.find(".") >= 0 or str.find("к") >= 0 or startWithDigit(str) == 1 or str.find("kk") >= 0 or str.find("k") >= 0:
                                if str.find('кг') == -1 and str.find('kg') == -1:
                                    if startWithDigit(str) == 1:
                                        if str.find('.') == 1 and str.find('.', 2) == 3 and str.find('.', 4) == 5:
                                            continue
                                        if str.find('.') == 1 or str.find(',') == 1: # проверка формата 5.1кк
                                            if str.find('кк') >= 0 or str.find('kk') >= 0: 
                                                #print('Выбран первый формат!')
                                                for str1 in str:
                                                    if block <= 2:
                                                        block = block + 1
                                                        if str1.isdigit() == 1 and bool == 0:
                                                            millions = int(str1)
                                                        if str1 == "." or str1 == ",":
                                                            bool = 1
                                                        if str1.isdigit() == 1 and bool == 1:
                                                            thousands = int(str1)
                                                millions = millions * 1000000
                                                thousands = thousands * 100000
                                                millions = millions + thousands
                                                return millions
                                        if str.find('.') == 3 or str.find('к') == 3 and str.find('.') != 1 or str.find('k') == 3 and str.find('.') != 1: # проверка формата 900к
                                            #print('Выбран второй формат!')
                                            count = 0
                                            for str1 in str:
                                                if count == 0:
                                                    thousands = int(str1)
                                                elif count == 1:
                                                    hundreds = int(str1)
                                                count = count + 1
                                            thousands = thousands * 100000
                                            hundreds = hundreds * 10000
                                            millions = thousands + hundreds
                                            return millions
                                        if str.find('.') == 1 and str.find('.', 2) == 5: # проверка формата  1.550.000
                                            #print('Выбран третий формат!')
                                            count = 0
                                            for str1 in str:
                                                if count <= 2:
                                                    if str1.isdigit() and count == 0:
                                                        millions = int(str1)
                                                        count = count + 1
                                                    elif str1.isdigit() and count == 1:
                                                        thousands = int(str1)
                                                        count = count + 1
                                                    elif str1.isdigit() and count == 2:
                                                        hundreds = int(str1)
                                                        count = count + 1
                                            millions = millions * 1000000
                                            thousands = thousands * 100000
                                            hundreds = hundreds * 10000
                                            millions = millions + thousands + hundreds
                                            return millions
                                        if str.find('.') == 1 or str.find(',') == 1: # проверка формата  1.550k
                                            if str.find('к') == 5 or str.find('k') == 5:
                                                #print('Выбран четвёртый формат!')
                                                count = 0
                                                for str1 in str:
                                                    if count <= 3:
                                                        if str1.isdigit() and count == 0:
                                                            millions = int(str1)
                                                            count = count + 1
                                                        elif str1.find('.') >= 0 and count == 1:
                                                            count = count + 1
                                                        elif str1.isdigit() and count == 2:
                                                            thousands = int(str1)
                                                            count = count + 1
                                                        elif str1.isdigit() and count == 3:
                                                            hundreds = int(str1)
                                                            count = count + 1
                                                millions = millions * 1000000
                                                thousands = thousands * 100000
                                                hundreds = hundreds * 10000
                                                millions = millions + thousands + hundreds
                                                return millions
                                        if str.find('кк') == 1 or str.find('kk') == 1: # проверка формата  1кк
                                            count = 0
                                            for str1 in str:
                                                if count == 0 and str1.isdigit() == 1:
                                                    millions = int(str1)
                                                    count = count + 1
                                            millions = millions * 1000000
                                            return millions

def checkPos(): # проверяет, что цена идёт после названия тачки
    None



def startWithDigit(str):
    if str.startswith('1') or str.startswith('2') or str.startswith('3') or str.startswith('4') or str.startswith('5') or str.startswith('6') or str.startswith('7') or str.startswith('8') or str.startswith('9'):
        return 1
    else:
        return 0

bot.run("mfa.68PoHXxzli3akZrhYYp-fa-CvS-P6PZvqC06WEq7w_FrEcAVbLPS-_UPLiimiDJr6NsqqgfGAI7LALDV0JjQ", bot=False)